"""
Excel Manager for handling all Excel file operations.
Provides CRUD operations for customer data stored in Excel format.
"""

import os
import openpyxl
from openpyxl import Workbook
from typing import List, Dict, Optional, Tuple
import pandas as pd
from .customer import Customer


class ExcelManager:
    """Manages Excel file operations for customer data."""
    
    def __init__(self, file_path: str = None):
        """Initialize Excel manager with file path."""
        # Use absolute path for better deployment compatibility
        if file_path:
            self.file_path = file_path
        else:
            # Try multiple possible locations for the Excel file
            possible_paths = [
                "business_excel.xlsx",
                "customers.xlsx", 
                os.path.join(os.path.dirname(__file__), "..", "business_excel.xlsx"),
                os.path.join(os.path.dirname(__file__), "..", "customers.xlsx"),
                os.path.join(os.getcwd(), "business_excel.xlsx"),
                os.path.join(os.getcwd(), "customers.xlsx")
            ]
            
            # Find the first existing file
            self.file_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    self.file_path = path
                    break
            
            # If no file exists, use the first option as default
            if not self.file_path:
                self.file_path = "business_excel.xlsx"
            
            # Debug: Print the selected file path
            print(f"ExcelManager: Using file path: {self.file_path}")
            print(f"ExcelManager: File exists: {os.path.exists(self.file_path)}")
            print(f"ExcelManager: Current working directory: {os.getcwd()}")
            print(f"ExcelManager: Files in current directory: {os.listdir('.')}")
        
        self.workbook = None
        self.worksheet = None
        self.headers = [
            'PHONE', 'CUSTOMER NAME', 'SCH   \nDEL', 'APT\nNO', 
            'ADDRESS', 'SUBURB', 'PC'
        ]
        self._ensure_file_exists()
    
    def _ensure_file_exists(self) -> None:
        """Create Excel file with headers if it doesn't exist."""
        if not os.path.exists(self.file_path):
            self._create_new_file()
        else:
            self._validate_file_structure()
    
    def _create_new_file(self) -> None:
        """Create a new Excel file with proper headers."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Customers"
        
        # Add headers
        for col, header in enumerate(self.headers, 1):
            self.worksheet.cell(row=1, column=col, value=header)
        
        self._save_workbook()
        print(f"Created new Excel file: {self.file_path}")
    
    def _validate_file_structure(self) -> None:
        """Validate that existing Excel file has correct structure."""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            
            # Check if headers exist
            first_row = [cell.value for cell in self.worksheet[1]]
            if first_row != self.headers:
                print(f"Warning: Excel file headers don't match expected format.")
                print(f"Expected: {self.headers}")
                print(f"Found: {first_row}")
                
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            self._create_new_file()
    
    def _save_workbook(self) -> None:
        """Save the workbook to file."""
        try:
            self.workbook.save(self.file_path)
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            raise
    
    def load_customers(self) -> List[Customer]:
        """Load all customers from Excel file."""
        try:
            print(f"Loading customers from: {self.file_path}")
            print(f"File exists: {os.path.exists(self.file_path)}")
            print(f"Current working directory: {os.getcwd()}")
            
            if not os.path.exists(self.file_path):
                print(f"ERROR: Excel file not found at {self.file_path}")
                print("Available files in current directory:")
                try:
                    for file in os.listdir('.'):
                        print(f"  - {file}")
                except Exception as e:
                    print(f"  Error listing directory: {e}")
                
                # Try to create a new file with sample data
                print("Creating new Excel file with sample data...")
                try:
                    self._create_new_file()
                    print(f"Created new Excel file at: {self.file_path}")
                except Exception as e:
                    print(f"Failed to create new Excel file: {e}")
                
                return []
            
            # Use pandas for better performance with large files
            df = pd.read_excel(self.file_path, header=None)  # Don't use first row as header
            print(f"Excel file loaded with {len(df)} rows and columns: {list(df.columns)}")
            
            # Find the header row (skip empty rows at the beginning)
            header_row_index = None
            for i in range(min(5, len(df))):  # Check first 5 rows for headers
                row = df.iloc[i]
                if any(pd.notna(val) and 'PHONE' in str(val) for val in row):
                    header_row_index = i
                    break
            
            if header_row_index is None:
                print("Could not find header row in Excel file")
                return []
            
            # Get the header row
            header_row = df.iloc[header_row_index]
            print(f"Header row found at index {header_row_index}: {header_row.tolist()}")
            
            # Map column indices to field names
            column_mapping = {}
            for i, header in enumerate(header_row):
                if pd.isna(header):
                    continue
                header_str = str(header).strip()
                if header_str == 'PHONE':
                    column_mapping['PHONE'] = i
                elif header_str == 'CUSTOMER NAME':
                    column_mapping['CUSTOMER NAME'] = i
                elif header_str == 'SCH   \nDEL':
                    column_mapping['SCH   \nDEL'] = i
                elif header_str == 'APT\nNO':
                    column_mapping['APT\nNO'] = i
                elif header_str == 'ADDRESS':
                    column_mapping['ADDRESS'] = i
                elif header_str == 'SUBURB':
                    column_mapping['SUBURB'] = i
                elif header_str == 'PC':
                    column_mapping['PC'] = i
            
            print(f"Column mapping: {column_mapping}")
            
            customers = []
            # Skip the header row and process data rows
            for index in range(header_row_index + 1, len(df)):
                try:
                    row = df.iloc[index]
                    
                    # Get values using column mapping
                    phone_raw = row.iloc[column_mapping.get('PHONE', -1)] if 'PHONE' in column_mapping else None
                    customer_name_raw = row.iloc[column_mapping.get('CUSTOMER NAME', -1)] if 'CUSTOMER NAME' in column_mapping else None
                    address_raw = row.iloc[column_mapping.get('ADDRESS', -1)] if 'ADDRESS' in column_mapping else None
                    
                    # Check if values are pandas NaN or None
                    if pd.isna(phone_raw) or str(phone_raw).strip() == '':
                        print(f"Skipping row {index}: Empty phone number")
                        continue
                    
                    if pd.isna(customer_name_raw) or str(customer_name_raw).strip() == '':
                        print(f"Skipping row {index}: Empty customer name")
                        continue
                    
                    if pd.isna(address_raw) or str(address_raw).strip() == '':
                        print(f"Skipping row {index}: Empty address")
                        continue
                    
                    # Clean the values - handle newlines and extra spaces
                    phone = str(phone_raw).strip().replace('\n', ' ').replace('\r', ' ')
                    customer_name = str(customer_name_raw).strip().replace('\n', ' ').replace('\r', ' ')
                    address = str(address_raw).strip().replace('\n', ' ').replace('\r', ' ')
                    
                    # Get other fields
                    sch_del_raw = row.iloc[column_mapping.get('SCH   \nDEL', -1)] if 'SCH   \nDEL' in column_mapping else ''
                    apt_no_raw = row.iloc[column_mapping.get('APT\nNO', -1)] if 'APT\nNO' in column_mapping else ''
                    suburb_raw = row.iloc[column_mapping.get('SUBURB', -1)] if 'SUBURB' in column_mapping else ''
                    pc_raw = row.iloc[column_mapping.get('PC', -1)] if 'PC' in column_mapping else ''
                    
                    # Clean other fields
                    sch_del = str(sch_del_raw).strip().replace('\n', ' ').replace('\r', ' ') if not pd.isna(sch_del_raw) else ''
                    apt_no = str(apt_no_raw).strip().replace('\n', ' ').replace('\r', ' ') if not pd.isna(apt_no_raw) else ''
                    suburb = str(suburb_raw).strip().replace('\n', ' ').replace('\r', ' ') if not pd.isna(suburb_raw) else ''
                    pc = str(pc_raw).strip().replace('\n', ' ').replace('\r', ' ') if not pd.isna(pc_raw) else ''
                    
                    # Only load rows with ALL required fields filled
                    customer_data = {
                        'PHONE': phone,
                        'CUSTOMER NAME': customer_name,
                        'SCH   \nDEL': sch_del,
                        'APT\nNO': apt_no,
                        'ADDRESS': address,
                        'SUBURB': suburb,
                        'PC': pc
                    }
                    
                    # Create customer object
                    customer = Customer.from_dict(customer_data)
                    customers.append(customer)
                    
                except Exception as e:
                    print(f"Error loading customer row {index}: {e}")
                    continue
            
            total_rows = len(df) - 1  # Subtract 1 for header row
            skipped_rows = total_rows - len(customers)
            print(f"Excel file had {total_rows} data rows (excluding header)")
            print(f"Skipped {skipped_rows} rows with missing required data")
            print(f"Successfully loaded {len(customers)} customers with complete data")
            return customers
            
        except Exception as e:
            print(f"Error loading customers from Excel: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def add_customer(self, customer: Customer) -> bool:
        """Add a new customer to Excel file."""
        try:
            # Load existing data using the same method as load_customers
            df = pd.read_excel(self.file_path, header=None)
            
            # Find the header row
            header_row_index = None
            for i in range(min(5, len(df))):
                row = df.iloc[i]
                if any(pd.notna(val) and 'PHONE' in str(val) for val in row):
                    header_row_index = i
                    break
            
            if header_row_index is None:
                print("Could not find header row for adding customer")
                return False
            
            # Get the header row to map columns
            header_row = df.iloc[header_row_index]
            column_mapping = {}
            for i, header in enumerate(header_row):
                if pd.isna(header):
                    continue
                header_str = str(header).strip()
                if header_str == 'PHONE':
                    column_mapping['PHONE'] = i
                elif header_str == 'CUSTOMER NAME':
                    column_mapping['CUSTOMER NAME'] = i
                elif header_str == 'SCH   \nDEL':
                    column_mapping['SCH   \nDEL'] = i
                elif header_str == 'APT\nNO':
                    column_mapping['APT\nNO'] = i
                elif header_str == 'ADDRESS':
                    column_mapping['ADDRESS'] = i
                elif header_str == 'SUBURB':
                    column_mapping['SUBURB'] = i
                elif header_str == 'PC':
                    column_mapping['PC'] = i
            
            # Create new row with proper column mapping
            new_row = [None] * len(df.columns)
            customer_dict = customer.to_dict()
            for col_name, col_index in column_mapping.items():
                if col_index < len(new_row):
                    new_row[col_index] = customer_dict.get(col_name, '')
            
            # Add the new row
            df.loc[len(df)] = new_row
            
            # Save back to Excel
            df.to_excel(self.file_path, index=False, header=False)
            
            print(f"Added customer: {customer.customer_name}")
            return True
            
        except Exception as e:
            print(f"Error adding customer: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def update_customer(self, old_customer: Customer, new_customer: Customer) -> bool:
        """Update an existing customer in Excel file."""
        try:
            # Load existing data using the same method as load_customers
            df = pd.read_excel(self.file_path, header=None)
            
            # Find the header row
            header_row_index = None
            for i in range(min(5, len(df))):
                row = df.iloc[i]
                if any(pd.notna(val) and 'PHONE' in str(val) for val in row):
                    header_row_index = i
                    break
            
            if header_row_index is None:
                print("Could not find header row for updating customer")
                return False
            
            # Get the header row to map columns
            header_row = df.iloc[header_row_index]
            column_mapping = {}
            for i, header in enumerate(header_row):
                if pd.isna(header):
                    continue
                header_str = str(header).strip()
                if header_str == 'PHONE':
                    column_mapping['PHONE'] = i
                elif header_str == 'CUSTOMER NAME':
                    column_mapping['CUSTOMER NAME'] = i
                elif header_str == 'SCH   \nDEL':
                    column_mapping['SCH   \nDEL'] = i
                elif header_str == 'APT\nNO':
                    column_mapping['APT\nNO'] = i
                elif header_str == 'ADDRESS':
                    column_mapping['ADDRESS'] = i
                elif header_str == 'SUBURB':
                    column_mapping['SUBURB'] = i
                elif header_str == 'PC':
                    column_mapping['PC'] = i
            
            # Clean the search data
            search_phone = str(old_customer.phone).strip().replace('\n', ' ').replace('\r', ' ') if old_customer.phone else ""
            search_name = str(old_customer.customer_name).strip().replace('\n', ' ').replace('\r', ' ') if old_customer.customer_name else ""
            search_address = str(old_customer.address).strip().replace('\n', ' ').replace('\r', ' ') if old_customer.address else ""
            
            print(f"DEBUG: Searching for - Phone: '{search_phone}', Name: '{search_name}', Address: '{search_address}'")
            
            # Find the customer row (skip header row)
            found_row = None
            for index in range(header_row_index + 1, len(df)):
                row = df.iloc[index]
                
                # Get values using column mapping
                phone_raw = row.iloc[column_mapping.get('PHONE', -1)] if 'PHONE' in column_mapping else None
                name_raw = row.iloc[column_mapping.get('CUSTOMER NAME', -1)] if 'CUSTOMER NAME' in column_mapping else None
                address_raw = row.iloc[column_mapping.get('ADDRESS', -1)] if 'ADDRESS' in column_mapping else None
                
                if pd.isna(phone_raw) or pd.isna(name_raw) or pd.isna(address_raw):
                    continue
                
                # Clean the values
                phone = str(phone_raw).strip().replace('\n', ' ').replace('\r', ' ')
                name = str(name_raw).strip().replace('\n', ' ').replace('\r', ' ')
                address = str(address_raw).strip().replace('\n', ' ').replace('\r', ' ')
                
                if phone == search_phone and name == search_name and address == search_address:
                    found_row = index
                    break
            
            if found_row is not None:
                # Update the row
                customer_dict = new_customer.to_dict()
                for col_name, col_index in column_mapping.items():
                    if col_index < len(df.columns):
                        df.iloc[found_row, col_index] = customer_dict.get(col_name, '')
                
                # Save back to Excel
                df.to_excel(self.file_path, index=False, header=False)
                
                print(f"Updated customer: {new_customer.customer_name}")
                return True
            else:
                print(f"Customer not found for update: {old_customer.customer_name}")
                return False
                
        except Exception as e:
            print(f"Error updating customer: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def delete_customer(self, customer: Customer) -> bool:
        """Delete a customer from Excel file."""
        try:
            # Load existing data using the same method as load_customers
            df = pd.read_excel(self.file_path, header=None)
            
            # Find the header row
            header_row_index = None
            for i in range(min(5, len(df))):
                row = df.iloc[i]
                if any(pd.notna(val) and 'PHONE' in str(val) for val in row):
                    header_row_index = i
                    break
            
            if header_row_index is None:
                print("Could not find header row for deleting customer")
                return False
            
            # Get the header row to map columns
            header_row = df.iloc[header_row_index]
            column_mapping = {}
            for i, header in enumerate(header_row):
                if pd.isna(header):
                    continue
                header_str = str(header).strip()
                if header_str == 'PHONE':
                    column_mapping['PHONE'] = i
                elif header_str == 'CUSTOMER NAME':
                    column_mapping['CUSTOMER NAME'] = i
                elif header_str == 'SCH   \nDEL':
                    column_mapping['SCH   \nDEL'] = i
                elif header_str == 'APT\nNO':
                    column_mapping['APT\nNO'] = i
                elif header_str == 'ADDRESS':
                    column_mapping['ADDRESS'] = i
                elif header_str == 'SUBURB':
                    column_mapping['SUBURB'] = i
                elif header_str == 'PC':
                    column_mapping['PC'] = i
            
            # Clean the search data
            search_phone = str(customer.phone).strip().replace('\n', ' ').replace('\r', ' ') if customer.phone else ""
            search_name = str(customer.customer_name).strip().replace('\n', ' ').replace('\r', ' ') if customer.customer_name else ""
            search_address = str(customer.address).strip().replace('\n', ' ').replace('\r', ' ') if customer.address else ""
            
            print(f"DEBUG: Searching for - Phone: '{search_phone}', Name: '{search_name}', Address: '{search_address}'")
            
            # Find the customer row (skip header row)
            found_row = None
            for index in range(header_row_index + 1, len(df)):
                row = df.iloc[index]
                
                # Get values using column mapping
                phone_raw = row.iloc[column_mapping.get('PHONE', -1)] if 'PHONE' in column_mapping else None
                name_raw = row.iloc[column_mapping.get('CUSTOMER NAME', -1)] if 'CUSTOMER NAME' in column_mapping else None
                address_raw = row.iloc[column_mapping.get('ADDRESS', -1)] if 'ADDRESS' in column_mapping else None
                
                if pd.isna(phone_raw) or pd.isna(name_raw) or pd.isna(address_raw):
                    continue
                
                # Clean the values
                phone = str(phone_raw).strip().replace('\n', ' ').replace('\r', ' ')
                name = str(name_raw).strip().replace('\n', ' ').replace('\r', ' ')
                address = str(address_raw).strip().replace('\n', ' ').replace('\r', ' ')
                
                if phone == search_phone and name == search_name and address == search_address:
                    found_row = index
                    break
            
            if found_row is not None:
                # Remove the row
                df = df.drop(found_row).reset_index(drop=True)
                
                # Save back to Excel
                df.to_excel(self.file_path, index=False, header=False)
                
                print(f"Deleted customer: {customer.customer_name}")
                return True
            else:
                print(f"Customer not found for deletion: {customer.customer_name}")
                return False
                
        except Exception as e:
            print(f"Error deleting customer: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def search_customers(self, query: str) -> List[Customer]:
        """Search customers by name, phone, suburb, or postal code."""
        try:
            customers = self.load_customers()
            query = query.lower().strip()
            
            if not query:
                return customers
            
            filtered_customers = []
            for customer in customers:
                if (query in customer.customer_name.lower() or
                    query in customer.phone.lower() or
                    query in customer.suburb.lower() or
                    query in customer.postal_code.lower()):
                    filtered_customers.append(customer)
            
            return filtered_customers
            
        except Exception as e:
            print(f"Error searching customers: {e}")
            return []
    
    def get_customer_by_phone(self, phone: str) -> Optional[Customer]:
        """Get customer by phone number."""
        try:
            customers = self.load_customers()
            for customer in customers:
                if customer.phone == phone:
                    return customer
            return None
        except Exception as e:
            print(f"Error getting customer by phone: {e}")
            return None
    
    def export_to_csv(self, output_path: str, customers: List[Customer] = None) -> bool:
        """Export customers to CSV file."""
        try:
            if customers is None:
                customers = self.load_customers()
            
            df = pd.DataFrame([customer.to_dict() for customer in customers])
            df.to_csv(output_path, index=False)
            
            print(f"Exported {len(customers)} customers to: {output_path}")
            return True
            
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False
    
    def get_statistics(self) -> Dict[str, int]:
        """Get basic statistics about the customer data."""
        try:
            customers = self.load_customers()
            
            stats = {
                'total_customers': len(customers),
                'customers_with_phone': len([c for c in customers if c.phone.strip()]),
                'customers_with_address': len([c for c in customers if c.address.strip()]),
                'customers_with_suburb': len([c for c in customers if c.suburb.strip()]),
                'customers_with_postal_code': len([c for c in customers if c.postal_code.strip()])
            }
            
            return stats
            
        except Exception as e:
            print(f"Error getting statistics: {e}")
            return {}
    
    def backup_file(self, backup_path: str = None) -> bool:
        """Create a backup of the Excel file."""
        try:
            if backup_path is None:
                import datetime
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = f"backup_{timestamp}_{os.path.basename(self.file_path)}"
            
            import shutil
            shutil.copy2(self.file_path, backup_path)
            
            print(f"Backup created: {backup_path}")
            return True
            
        except Exception as e:
            print(f"Error creating backup: {e}")
            return False
