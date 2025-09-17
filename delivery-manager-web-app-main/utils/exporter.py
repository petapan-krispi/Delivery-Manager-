"""
Data exporter utility for saving customer data in various formats.
Supports CSV, Excel, and TXT export for web app.
"""

import csv
import os
import io
from typing import List, Dict, Any
from models.customer import Customer
from datetime import datetime


class DataExporter:
    """Handles exporting customer data to various file formats for web app."""
    
    def __init__(self):
        """Initialize the data exporter."""
        self.supported_formats = ['csv', 'txt', 'xlsx']
    
    def export_to_csv(self, customers: List[Customer]) -> bytes:
        """Export customers to CSV format and return as bytes."""
        try:
            if not customers:
                return b""
            
            # Create CSV in memory
            output = io.StringIO()
            
            # Define CSV headers
            headers = [
                'Phone',
                'Customer Name', 
                'Scheduled Delivery Time',
                'Apartment No',
                'Address',
                'Suburb',
                'Postal Code',
                'Export Date'
            ]
            
            # Write CSV
            writer = csv.writer(output)
            writer.writerow(headers)
            
            # Write customer data
            export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for customer in customers:
                row = [
                    customer.phone,
                    customer.customer_name,
                    customer.scheduled_delivery_time,
                    customer.apartment_no,
                    customer.address,
                    customer.suburb,
                    customer.postal_code,
                    export_date
                ]
                writer.writerow(row)
            
            # Get CSV content as bytes
            csv_content = output.getvalue()
            output.close()
            
            return csv_content.encode('utf-8')
            
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return b""
    
    def export_to_excel(self, customers: List[Customer]) -> bytes:
        """Export customers to Excel format and return as bytes."""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            if not customers:
                return b""
            
            # Create workbook and worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Customers"
            
            # Define headers
            headers = [
                'Phone',
                'Customer Name', 
                'Scheduled Delivery Time',
                'Apartment No',
                'Address',
                'Suburb',
                'Postal Code',
                'Export Date'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Write customer data
            export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for row, customer in enumerate(customers, 2):
                worksheet.cell(row=row, column=1, value=customer.phone)
                worksheet.cell(row=row, column=2, value=customer.customer_name)
                worksheet.cell(row=row, column=3, value=customer.scheduled_delivery_time)
                worksheet.cell(row=row, column=4, value=customer.apartment_no)
                worksheet.cell(row=row, column=5, value=customer.address)
                worksheet.cell(row=row, column=6, value=customer.suburb)
                worksheet.cell(row=row, column=7, value=customer.postal_code)
                worksheet.cell(row=row, column=8, value=export_date)
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return b""
    
    def export_to_txt(self, customers: List[Customer]) -> bytes:
        """Export customers to plain text format and return as bytes."""
        try:
            if not customers:
                return b""
            
            # Create text content
            output = io.StringIO()
            
            # Write header
            output.write("DELIVERY MANAGER - CUSTOMER EXPORT\n")
            output.write("=" * 50 + "\n")
            output.write(f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            output.write(f"Total Customers: {len(customers)}\n")
            output.write("=" * 50 + "\n\n")
            
            # Write customer data
            for i, customer in enumerate(customers, 1):
                output.write(f"CUSTOMER {i}\n")
                output.write("-" * 20 + "\n")
                output.write(f"Name: {customer.customer_name}\n")
                output.write(f"Phone: {customer.phone}\n")
                output.write(f"Address: {customer.address}\n")
                
                if customer.apartment_no:
                    output.write(f"Apartment: {customer.apartment_no}\n")
                
                output.write(f"Suburb: {customer.suburb}\n")
                output.write(f"Postal Code: {customer.postal_code}\n")
                
                if customer.scheduled_delivery_time:
                    output.write(f"Scheduled Delivery: {customer.scheduled_delivery_time}\n")
                
                output.write("\n")
            
            # Get text content as bytes
            text_content = output.getvalue()
            output.close()
            
            return text_content.encode('utf-8')
            
        except Exception as e:
            print(f"Error exporting to TXT: {e}")
            return b""
    
    def get_export_filename(self, format_type: str, search_query: str = "") -> str:
        """Get filename for export based on format and search query."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        query_part = f"_{search_query.replace(' ', '_')}" if search_query else ""
        
        if format_type == 'csv':
            return f"customers{query_part}_{timestamp}.csv"
        elif format_type == 'xlsx':
            return f"customers{query_part}_{timestamp}.xlsx"
        elif format_type == 'txt':
            return f"customers{query_part}_{timestamp}.txt"
        else:
            return f"export_{timestamp}.{format_type}"
    
    def get_export_mime_type(self, format_type: str) -> str:
        """Get MIME type for export format."""
        mime_types = {
            'csv': 'text/csv',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'txt': 'text/plain'
        }
        return mime_types.get(format_type, 'application/octet-stream')
    
    def get_export_summary(self, customers: List[Customer]) -> Dict[str, Any]:
        """Get summary information for export."""
        try:
            if not customers:
                return {
                    'total_customers': 0,
                    'export_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'file_size_estimate': 0
                }
            
            # Calculate estimated file size (rough estimate)
            avg_customer_size = 200  # bytes per customer (rough estimate)
            estimated_size = len(customers) * avg_customer_size
            
            return {
                'total_customers': len(customers),
                'export_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'file_size_estimate': estimated_size,
                'fields_count': 8,  # Number of fields being exported
                'has_apartment_data': any(c.apartment_no for c in customers),
                'has_delivery_time': any(c.scheduled_delivery_time for c in customers)
            }
            
        except Exception as e:
            print(f"Error generating export summary: {e}")
            return {}
